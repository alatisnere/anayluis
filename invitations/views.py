import csv

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .forms import AppConfigForm, PartyCreateForm, PartyCSVUploadForm, RSVPForm
from .models import AppConfig, Party, RSVP
from .services import build_wa_me_link, compute_status

SESSION_KEY = "invite_access_granted"


def dashboard_allowed(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff)


def rsvp_detail(request, code):
    party = get_object_or_404(Party, code=code)

    config = AppConfig.get_solo()
    if not config:
        return redirect("app_config")

    now = timezone.now()
    deadline_passed = now > config.rsvp_deadline

    existing = getattr(party, "rsvp", None)
    edit_mode = request.GET.get("edit") == "1"

    initial = {}
    if existing:
        initial = {
            "response": existing.response,
            "tickets_confirmed": existing.tickets_confirmed,
        }

    if request.method == "POST":
        if deadline_passed:
            messages.error(request, "❌ El periodo de confirmación ya terminó.")
            return redirect("rsvp_detail", code=str(party.code))

        form = RSVPForm(request.POST, max_tickets=party.max_tickets)
        if form.is_valid():
            data = form.cleaned_data

            rsvp, _ = RSVP.objects.update_or_create(
                party=party,
                defaults={
                    "response": data["response"],
                    "tickets_confirmed": data["tickets_confirmed"],
                },
            )

            if rsvp.response == RSVP.Response.NO:
                messages.success(request, config.confirm_text_no)
            elif rsvp.tickets_confirmed >= party.max_tickets:
                messages.success(request, config.confirm_text_yes)
            else:
                messages.success(request, config.confirm_text_partial)

            return redirect("rsvp_detail", code=str(party.code))
    else:
        form = RSVPForm(initial=initial, max_tickets=party.max_tickets)

    show_summary = bool(existing) and not edit_mode
    confirmed_tickets = existing.tickets_confirmed if existing else 0

    deadline_message = None
    if deadline_passed:
        if not existing:
            deadline_message = config.rsvp_deadline_passed_text_none
        elif existing.response == RSVP.Response.NO:
            deadline_message = config.rsvp_deadline_passed_text_no
        elif existing.tickets_confirmed >= party.max_tickets:
            deadline_message = config.rsvp_deadline_passed_text_yes
        else:
            deadline_message = config.rsvp_deadline_passed_text_partial

    return render(
        request,
        "invitations/rsvp_detail.html",
        {
            "party": party,
            "form": form,
            "rsvp": existing,
            "confirmed_tickets": confirmed_tickets,
            "show_summary": show_summary,
            "edit_mode": edit_mode,
            "config": config,
            "deadline_passed": deadline_passed,
            "deadline_message": deadline_message,
        },
    )


@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.user.is_authenticated:
        if dashboard_allowed(request.user):
            return redirect("dashboard")
        logout(request)
        return render(
            request,
            "auth/login.html",
            {"error": "No tienes permisos para acceder al dashboard."},
        )

    error = None
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)

        if user is None:
            error = "Invalid credentials."
        elif not (user.is_superuser or user.is_staff):
            error = "No tienes permisos para acceder al dashboard."
        else:
            login(request, user)
            return redirect("dashboard")

    return render(request, "auth/login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("login")


@user_passes_test(dashboard_allowed, login_url="/login/")
def dashboard(request):
    config = AppConfig.get_solo()
    if not config:
        return redirect("app_config")

    now = timezone.now()
    deadline = config.rsvp_deadline
    delta = deadline - now

    if delta.total_seconds() <= 0:
        rsvp_countdown_text = "El RSVP ya se cerró"
    else:
        total_seconds = int(delta.total_seconds())
        if total_seconds < 3600:
            minutes = max(1, total_seconds // 60)
            rsvp_countdown_text = f"El RSVP se cierra en {minutes} min"
        elif total_seconds < 86400:
            hours = max(1, total_seconds // 3600)
            rsvp_countdown_text = f"El RSVP se cierra en {hours} hora{'s' if hours != 1 else ''}"
        else:
            days = max(1, total_seconds // 86400)
            rsvp_countdown_text = f"El RSVP se cierra en {days} día{'s' if days != 1 else ''}"

    rsvp_deadline_date = deadline

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    per_page_choices = [10, 20, 30, 40, 50]
    try:
        per_page = int(request.GET.get("per_page") or 10)
    except ValueError:
        per_page = 10
    if per_page not in per_page_choices:
        per_page = 10

    try:
        page_number = int(request.GET.get("page") or 1)
    except ValueError:
        page_number = 1

    all_parties_qs = Party.objects.select_related("rsvp").all()

    family_counts = {
        "TOTAL": 0,
        "ACCEPTED": 0,
        "PARTIAL": 0,
        "DECLINED": 0,
        "NO_RESPONSE": 0,
    }

    ticket_counts = {
        "ASSIGNED": 0,
        "ACCEPTED": 0,
        "DENIED": 0,
        "PENDING": 0,
    }

    for p in all_parties_qs:
        family_counts["TOTAL"] += 1
        ticket_counts["ASSIGNED"] += p.max_tickets

        status = compute_status(p)
        family_counts[status.key] += 1

        rsvp = getattr(p, "rsvp", None)
        confirmed = int(getattr(rsvp, "tickets_confirmed", 0) or 0)

        if status.key == "NO_RESPONSE":
            ticket_counts["PENDING"] += p.max_tickets
        else:
            ticket_counts["ACCEPTED"] += confirmed
            denied = p.max_tickets - confirmed
            if denied > 0:
                ticket_counts["DENIED"] += denied

    table_qs = all_parties_qs.order_by("display_name")
    if q:
        table_qs = table_qs.filter(display_name__icontains=q)

    filtered_parties = []
    for p in table_qs:
        status = compute_status(p)
        if status_filter and status.key != status_filter:
            continue
        filtered_parties.append((p, status))

    rows = []
    for p, status in filtered_parties:
        rsvp = getattr(p, "rsvp", None)
        confirmed = int(getattr(rsvp, "tickets_confirmed", 0) or 0)

        rsvp_url = request.build_absolute_uri(reverse("rsvp_detail", args=[str(p.code)]))
        template = config.invite_message_template or "{link}"
        generated_text = template.replace("{link}", rsvp_url)
        wa_link = build_wa_me_link(p.phone_e164, generated_text)

        rows.append(
            {
                "party": p,
                "status": status,
                "tickets_confirmed": confirmed if rsvp else None,
                "created_at": getattr(rsvp, "created_at", None),
                "updated_at": getattr(rsvp, "updated_at", None),
                "rsvp_url": rsvp_url,
                "wa_text": generated_text,
                "wa_link": wa_link,
            }
        )

    paginator = Paginator(rows, per_page)
    page_obj = paginator.get_page(page_number)

    statuses = [
        ("ACCEPTED", "Aceptaron"),
        ("PARTIAL", "Parcial"),
        ("DECLINED", "Denegaron"),
        ("NO_RESPONSE", "Pendiente"),
    ]

    return render(
        request,
        "dashboard/dashboard.html",
        {
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "per_page": per_page,
            "per_page_choices": per_page_choices,
            "family_counts": family_counts,
            "ticket_counts": ticket_counts,
            "q": q,
            "status_filter": status_filter,
            "statuses": statuses,
            "config": config,
            "rsvp_countdown_text": rsvp_countdown_text,
            "rsvp_deadline_date": rsvp_deadline_date,
        },
    )


@user_passes_test(dashboard_allowed, login_url="/login/")
def export_csv(request):
    parties = Party.objects.select_related("rsvp").all().order_by("display_name")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="rsvp_bodas.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)

    writer.writerow(
        [
            "familia",
            "boletos_asignados",
            "telefono_whatsapp",
            "mostrar_texto_restricciones",
            "estado_familia",
            "boletos_confirmados",
            "fecha_respuesta",
            "fecha_actualizacion",
        ]
    )

    def yn(value: bool) -> str:
        return "sí" if value else "no"

    def status_es(key: str) -> str:
        return {
            "ACCEPTED": "Aceptaron",
            "PARTIAL": "Parcial",
            "DECLINED": "Denegaron",
            "NO_RESPONSE": "Pendiente",
        }.get(key, key)

    def fmt_dt(dt):
        if not dt:
            return ""
        dt = timezone.localtime(dt)
        return dt.strftime("%d/%m/%Y %H:%M")

    for p in parties:
        rsvp = getattr(p, "rsvp", None)
        status = compute_status(p)

        writer.writerow(
            [
                p.display_name,
                p.max_tickets,
                p.phone_e164 or "",
                yn(bool(p.show_restrictions_text)),
                status_es(status.key),
                getattr(rsvp, "tickets_confirmed", "") if rsvp else "",
                fmt_dt(getattr(rsvp, "created_at", None)),
                fmt_dt(getattr(rsvp, "updated_at", None)),
            ]
        )

    return response


@require_http_methods(["GET", "POST"])
def gate(request):
    if settings.DEMO_MODE or request.session.get(SESSION_KEY):
        request.session[SESSION_KEY] = True
        return redirect("home")

    config = AppConfig.get_solo()
    if not config:
        return render(request, "public/password.html", {"configured": False})

    if request.method == "POST":
        pw = (request.POST.get("password", "") or "").strip()
        if pw == config.invite_password:
            request.session[SESSION_KEY] = True
            return redirect("home")
        return render(
            request,
            "public/password.html",
            {"configured": True, "error": "Contraseña incorrecta."},
        )

    return render(request, "public/password.html", {"configured": True})


def home(request):
    config = AppConfig.get_solo()
    if not config:
        return redirect("acceso")

    if settings.DEMO_MODE:
        request.session[SESSION_KEY] = True
        return render(request, "public/home.html")

    url_pw = request.GET.get("password")
    if url_pw and url_pw == config.invite_password:
        request.session[SESSION_KEY] = True
        return render(request, "public/home.html")

    if not request.session.get(SESSION_KEY):
        return redirect("acceso")

    return render(request, "public/home.html")


def logout_invite(request):
    request.session.pop(SESSION_KEY, None)
    return redirect("gate")


@user_passes_test(dashboard_allowed, login_url="/login/")
@require_http_methods(["POST"])
def party_create(request):
    form = PartyCreateForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "✅ Familia creada.")
    else:
        messages.error(request, "❌ Revisa los campos del formulario.")
    return redirect("dashboard")


@user_passes_test(dashboard_allowed, login_url="/login/")
@require_http_methods(["POST"])
def party_update(request, party_id):
    party = get_object_or_404(Party, id=party_id)

    form = PartyCreateForm(request.POST, instance=party)
    if form.is_valid():
        form.save()
        messages.success(request, "✅ Familia actualizada.")
    else:
        messages.error(request, "❌ Revisa los campos del formulario.")

    return redirect("dashboard")


@user_passes_test(dashboard_allowed, login_url="/login/")
@require_http_methods(["GET", "POST"])
def party_upload_csv(request):
    """
    Bulk import Parties from CSV.

    Expected headers shown to the user:
      - nombre_invitado (required)  -> Party.display_name
      - tickets_asignados (required)-> Party.max_tickets
      - whatsapp (optional)         -> Party.phone_e164
      - texto_restricciones (optional; si/no, true/false, 1/0, etc.) -> Party.show_restrictions_text (bool)
    """
    if request.method == "POST":
        form = PartyCSVUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "❌ Sube un archivo CSV.")
            return redirect("party_upload_csv")

        f = form.cleaned_data["file"]
        created = 0
        skipped = 0

        try:
            content = f.read().decode("utf-8-sig").splitlines()
        except Exception:
            messages.error(request, "❌ No se pudo leer el archivo. Sube un CSV válido en UTF-8.")
            return redirect("party_upload_csv")

        reader = csv.DictReader(content)
        fieldnames = set((reader.fieldnames or []))

        required_cols = {"nombre_invitado", "tickets_asignados"}
        if not required_cols.issubset(fieldnames):
            messages.error(
                request,
                "❌ El archivo debe incluir columnas: "
                "nombre_invitado, tickets_asignados. "
                "Opcional: whatsapp, texto_restricciones.",
            )
            return redirect("party_upload_csv")

        def parse_si_no(val) -> bool:
            v = (val or "").strip().lower()
            if v in {"si", "sí", "s", "y", "yes", "true", "t", "1", "x"}:
                return True
            if v in {"no", "n", "false", "f", "0"}:
                return False
            return False

        with transaction.atomic():
            for row in reader:
                name = (row.get("nombre_invitado") or "").strip()
                tickets_raw = (row.get("tickets_asignados") or "").strip()
                phone = (row.get("whatsapp") or "").strip()
                show_restrictions = parse_si_no(row.get("texto_restricciones"))

                if not name or not tickets_raw:
                    skipped += 1
                    continue

                try:
                    max_tickets = int(tickets_raw)
                    if max_tickets < 1:
                        skipped += 1
                        continue
                except ValueError:
                    skipped += 1
                    continue

                Party.objects.create(
                    display_name=name,
                    max_tickets=max_tickets,
                    phone_e164=phone,
                    show_restrictions_text=show_restrictions,
                )
                created += 1

        messages.success(request, f"CSV procesado. Creadas: {created}. Omitidas: {skipped}.")
        return redirect("dashboard")

    form = PartyCSVUploadForm()
    return render(request, "dashboard/upload_csv.html", {"form": form})


@user_passes_test(dashboard_allowed, login_url="/login/")
def party_template_xlsx(request):
    """
    XLSX template with user-facing headers only:
      nombre_invitado, tickets_asignados, whatsapp, texto_restricciones
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invitados"

    headers = ["nombre_invitado", "tickets_asignados", "whatsapp", "texto_restricciones"]
    ws.append(headers)
    ws.append(["Familia de ejemplo", 4, "5212220000000", "si"])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(h) + 2)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_invitados.xlsx"'
    wb.save(response)
    return response


@user_passes_test(dashboard_allowed, login_url="/login/")
@require_POST
def party_delete(request, party_id):
    party = get_object_or_404(Party, id=party_id)
    party.delete()
    messages.success(request, "✅ Invitado eliminado.")
    return redirect("dashboard")


@user_passes_test(dashboard_allowed, login_url="/login/")
@require_http_methods(["GET", "POST"])
def app_config(request):
    config = AppConfig.get_solo()

    if request.method == "POST":
        form = AppConfigForm(request.POST, instance=config)
        if form.is_valid():
            obj = form.save(commit=False)
            if config is None:
                obj.pk = 1
            obj.save()
            messages.success(request, "✅ Configuración guardada.")
            return redirect("dashboard")
        messages.error(request, "❌ Revisa los campos.")
    else:
        form = AppConfigForm(instance=config)

    return render(
        request,
        "dashboard/config.html",
        {
            "form": form,
            "config": config,
        },
    )
